'''task-1'''
from openpyxl import load_workbook

if __name__ == "__main__":

    wb2 = load_workbook('test.xlsx')
    print(wb2.sheetnames)

    wb = load_workbook('test.xlsx')
    ws = wb.active
    print(ws["A1"].value)
    print(ws.cell(row=1, column = 1, value = 44))
    print(ws.cell( row = 1, column = 1).value)